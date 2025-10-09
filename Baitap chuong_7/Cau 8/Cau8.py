from openpyxl import load_workbook
wb = load_workbook('Baitap chuong_7/Cau 7/demo.xlsx')
print (wb.sheetnames)
ws = wb[wb.sheetnames[0]]
for row in ws.values:
    for value in row:
        print(value,"\t",end='')
    print("")